import logging
import re
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# کتابخانه‌های مورد نیاز برای تلگرام
from telegram import Update, ForceReply
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    ConversationHandler,
    ContextTypes,
)

# --- Configuration ---
# --- تنظیمات اصلی ربات ---
# توکن ربات تلگرام شما
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "8265715246:AAGvKvPBhbFry2ARrWXZDye7vHfmLfPgHZY")
# نام فایل اکسل که اطلاعات در آن ذخیره می‌شود
XLSX_FILENAME = "bot_data.xlsx"
# لیستی از شناسه‌های کاربری که باید بلاک شوند
# می‌توانید هر تعداد شناسه که خواستید به این لیست اضافه کنید
BLOCKED_USER_IDS = {6779565731,1362772795,52316973}


# --- Logging Setup ---
# فعال‌سازی لاگ‌ها برای اینکه بتوانید فعالیت‌ها و خطاهای ربات را ببینید
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO
)
logger = logging.getLogger(__name__)

# --- Conversation States ---
# تعریف وضعیت‌های مختلف مکالمه برای مدیریت مراحل دریافت اطلاعات
GET_NAME, GET_EMAIL = range(2)

# --- Excel File Setup ---
def setup_excel_file():
    """
    اگر فایل اکسل وجود نداشته باشد با یک شیت و هدرهای مناسب ایجاد می‌شود.
    """
    if not os.path.exists(XLSX_FILENAME):
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Data"
            headers = ["Timestamp", "Name", "Email", "User ID"]
            sheet.append(headers)
            workbook.save(XLSX_FILENAME)
            logger.info(f"Excel file '{XLSX_FILENAME}' created with headers.")
        except Exception as e:
            logger.error(f"Failed to create Excel file: {e}")

def append_row_to_excel(row_values: list) -> None:
    """افزودن یک ردیف به فایل اکسل."""
    workbook = load_workbook(XLSX_FILENAME)
    sheet = workbook.active
    sheet.append(row_values)
    workbook.save(XLSX_FILENAME)


# --- Helper Functions ---
def is_valid_email(email: str) -> bool:
    """Check if the provided string is a valid email address using regex."""
    regex = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    return re.match(regex, email) is not None

# --- Telegram Bot Handlers ---
# --- توابع مدیریت‌کننده دستورات و پیام‌های تلگرام ---

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """شروع مکالمه، بررسی کاربر بلاک شده و درخواست نام"""
    user = update.effective_user

    # --- بخش جدید: بررسی شناسه کاربر در لیست بلاک ---
    if user.id in BLOCKED_USER_IDS:
        logger.warning(f"Blocked user {user.id} ({user.full_name}) tried to start the bot.")
        # هیچ پیامی به کاربر بلاک شده ارسال نمی‌شود
        return ConversationHandler.END # مکالمه را فوراً و در سکوت پایان می‌دهد

    # اگر کاربر بلاک نباشد، ربات به کار عادی خود ادامه می‌دهد
    logger.info(f"User {user.id} ({user.full_name}) started the bot.")
    await update.message.reply_html(
        f"سلام {user.mention_html()}! 👋\n"
        "برای ذخیره اطلاعاتت، لطفاً اسمت رو وارد کن."
    )
    return GET_NAME

async def get_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """ذخیره نام و درخواست ایمیل"""
    user_name = update.message.text
    context.user_data['name'] = user_name
    logger.info(f"User {update.effective_user.id} entered name: {user_name}")
    await update.message.reply_text(
        "ممنون! حالا لطفا ایمیل‌تون رو وارد کنید.",
        reply_markup=ForceReply(selective=True),
    )
    return GET_EMAIL

async def get_email(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """ذخیره ایمیل، اعتبارسنجی و ثبت نهایی اطلاعات در فایل اکسل"""
    user_email = update.message.text
    user_id = update.effective_user.id
    user_name = context.user_data.get('name', 'N/A')
    timestamp = datetime.now().isoformat()

    logger.info(f"User {user_id} entered email: {user_email}")

    if not is_valid_email(user_email):
        await update.message.reply_text(
            "ایمیل وارد شده معتبر نیست. لطفاً یک ایمیل صحیح وارد کنید."
        )
        return GET_EMAIL

    try:
        row = [timestamp, user_name, user_email, str(user_id)]
        append_row_to_excel(row)
        logger.info(f"Data saved to Excel for user {user_id}: {row}")
        await update.message.reply_text(
            "اطلاعات شما با موفقیت ذخیره شد. ممنون از همکاری‌تون! 😊"
        )
    except Exception as e:
        logger.error(f"Error saving data to Excel for user {user_id}: {e}")
        await update.message.reply_text(
            "متاسفانه مشکلی در ذخیره اطلاعات پیش آمد. لطفاً دوباره تلاش کنید."
        )
    
    context.user_data.clear()
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """لغو مکالمه در هر مرحله"""
    user = update.effective_user
    logger.info(f"User {user.id} canceled the conversation.")
    await update.message.reply_text(
        "مکالمه لغو شد. هر وقت خواستید دوباره شروع کنید، /start رو بزنید."
    )
    context.user_data.clear()
    return ConversationHandler.END

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """مدیریت خطاهای پیش‌بینی نشده"""
    logger.error(msg="Exception while handling an update:", exc_info=context.error)
    if isinstance(update, Update) and update.effective_message:
        await update.effective_message.reply_text(
            "متاسفانه یک خطای ناشناخته رخ داد. لطفاً دوباره تلاش کنید."
        )

def main() -> None:
    """تابع اصلی برای راه‌اندازی و اجرای ربات"""
    
    setup_excel_file()
    
    application = Application.builder().token(TELEGRAM_BOT_TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            GET_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_name)],
            GET_EMAIL: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_email)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )

    application.add_handler(conv_handler)
    application.add_error_handler(error_handler)

    logger.info("Bot is running...")
    application.run_polling(drop_pending_updates=True)

if __name__ == '__main__':
    main()


